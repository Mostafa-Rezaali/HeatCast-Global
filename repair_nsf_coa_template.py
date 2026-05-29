from copy import copy
from pathlib import Path

from openpyxl import load_workbook


OUT_DIR = Path(r"C:\Users\mosta\OneDrive\Desktop\GraphCast\nsf_ags_prf_uploads\personnel_latex")
TEMPLATE = OUT_DIR / "coa_template_official.xlsx"
OUTPUT = OUT_DIR / "collaborators_other_affiliations_mostafa_rezaali_fixed.xlsx"


def copy_row_style(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def clear_cell(cell):
    cell.value = None
    cell.hyperlink = None
    cell.comment = None


wb = load_workbook(TEMPLATE)
ws = wb["NSF COA Template"]

# Add enough Table 4 rows while preserving the official template's visual style.
table4_entries = [
    ["A:", "Narayanan, A.", None, None, None],
    ["A:", "Bunting, E. L.", None, None, None],
    ["A:", "Keellings, David", "University of Florida", "Department of Geography", None],
    ["A:", "Fouladi-Fard, R.", None, None, None],
    ["A:", "O'Shaughnessy, P.", None, None, None],
    ["A:", "Naddafi, K.", None, None, None],
    ["A:", "Karimi, A.", "Qom University of Technology", None, None],
    ["A:", "Jahangir, M. S.", None, None, None],
    ["A:", "Quilty, J.", None, None, None],
    ["A:", "Mojarad, H.", None, None, None],
    ["A:", "Sorooshian, A.", None, None, None],
    ["A:", "Mahdinia, M.", None, None, None],
    ["A:", "Farajollahi, M.", None, None, None],
    ["A:", "Fahiminia, M.", None, None, None],
    ["A:", "Rahimi, N. R.", None, None, None],
    ["A:", "Aali, R.", None, None, None],
    ["A:", "Shahryari, A.", None, None, None],
    ["C:", "Li, Shawn", "Columbia University", "Research project supervisor listed in CV", None],
]

extra_rows = max(0, len(table4_entries) - 6)
if extra_rows:
    ws.insert_rows(58, amount=extra_rows)
    for row in range(58, 58 + extra_rows):
        copy_row_style(ws, 57 + extra_rows, row)

# Clear the official sample values without changing instructions or format.
for row in [17, 18, 19, 28, 38, 39, 52, 53]:
    for col in range(1, 6):
        clear_cell(ws.cell(row, col))

# Table 1: current/near-current affiliations.
table1_rows = [
    [None, "Rezaali, Mostafa", "University of Florida", None, None],
    [None, None, "University of Florida, Department of Geography (Ph.D. candidate)", None, None],
]
for offset, values in enumerate(table1_rows, start=17):
    for col, value in enumerate(values, start=1):
        ws.cell(offset, col).value = value

# Table 2: no personal/family/business relationship identified from available materials.
for col in range(1, 6):
    ws.cell(28, col).value = None

# Table 3: Ph.D. and graduate advisors from available CV/application context.
advisor_rows = [
    ["G:", "Keellings, David", "University of Florida", "Department of Geography", None],
    ["G:", "Karimi, Ali", "Qom University of Technology", None, None],
]
for offset, values in enumerate(advisor_rows, start=38):
    for col, value in enumerate(values, start=1):
        clear_cell(ws.cell(offset, col))
        ws.cell(offset, col).value = value

# Table 4: coauthors and recent project collaborators.
for offset, values in enumerate(table4_entries, start=52):
    for col, value in enumerate(values, start=1):
        clear_cell(ws.cell(offset, col))
        ws.cell(offset, col).value = value
    copy_row_style(ws, 52 if values[0] == "A:" else 53, offset)

# Table 5: no editorial-board/co-editor relationships identified from available materials.
table5_start = 58 + extra_rows
ws.row_dimensions[table5_start].height = 29.1
for row in range(table5_start + 6, table5_start + 11):
    for col in range(1, 6):
        clear_cell(ws.cell(row, col))

# Keep all visible cells at the template's font size and wrapping behavior.
for row in range(17, 70 + extra_rows):
    for col in range(1, 6):
        ws.cell(row, col).alignment = copy(ws.cell(row, col).alignment)

wb.save(OUTPUT)
print(OUTPUT)
